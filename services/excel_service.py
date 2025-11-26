# services/excel_service.py
"""
Servicio para procesamiento de planillas Excel
Lee, procesa, valida y genera archivos Excel con datos del sistema
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Any, Optional, Union
from datetime import datetime, date
import io
import os

class ExcelService:
"""Servicio para manejar archivos Excel"""

def __init__(self, db_session=None):
self.db = db_session

# Configuración de estilos
self.header_font = Font(bold=True, color="FFFFFF")
self.header_fill = PatternFill(start_color="36609", end_color="36609", fill_type="solid")
self.border = Border(
left=Side(style='thin'),
right=Side(style='thin'),
top=Side(style='thin'),
bottom=Side(style='thin')
)
self.center_alignment = Alignment(horizontal='center', vertical='center')

def read_excel_file(self, file_path: str, sheet_name: str = None) -> Dict[str, Any]:
"""
Lee un archivo Excel y retorna los datos
"""
try:
# Leer archivo
if sheet_name:
df = pd.read_excel(file_path, sheet_name=sheet_name)
else:
# Leer primera hoja por defecto
df = pd.read_excel(file_path)

# Validar datos
validation = self._validate_dataframe(df)

if not validation['valid']:
return validation

# Procesar datos
processed_data = self._process_dataframe(df)

return {
'success': True,
'data': processed_data.to_dict('records'),
'columns': list(processed_data.columns),
'shape': processed_data.shape,
'sheet_names': [sheet_name] if sheet_name else ['Sheet1']
}

except Exception as e:
return {
'success': False,
'error': f'Error leyendo archivo Excel: {str(e)}'
}

def _validate_dataframe(self, df: pd.DataFrame) -> Dict[str, Any]:
"""
Valida un DataFrame de pandas
"""
if df.empty:
return {
'valid': False,
'error': 'El archivo está vacío'
}

if df.shape[0] == 0:
return {
'valid': False,
'error': 'No hay datos en la hoja'
}

# Verificar columnas requeridas
required_columns = []
return {'valid': True}

def _process_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
"""
Procesa y limpia un DataFrame
"""
# Limpiar espacios en blanco
df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

# Convertir fechas
date_columns = ['fecha', 'Fecha', 'DATE', 'created_at', 'updated_at']
for col in date_columns:
if col in df.columns:
df[col] = pd.to_datetime(df[col], errors='coerce')

# Convertir números
numeric_columns = ['ip', 'puerto', 'capacidad', 'autonomia', 'voltaje', 'corriente']
for col in numeric_columns:
if col in df.columns:
df[col] = pd.to_numeric(df[col], errors='coerce')

return df

def export_cameras_to_excel(self, filters: Dict = None) -> Dict[str, Any]:
"""
Exporta datos de cámaras a Excel
"""
try:
# Construir consulta
query = """
SELECT
c.id,
c.nombre,
c.tipo,
c.ip,
c.marca,
c.modelo,
c.estado,
c.fecha_instalacion,
l.nombre as ubicacion,
s.nombre as switch_conectado,
s.ip as switch_ip,
c.observaciones
FROM camaras c
LEFT JOIN ubicaciones l ON c.ubicacion_id = l.id
LEFT JOIN switches s ON c.switch_id = s.id
"""

params = []
conditions = []

# Aplicar filtros
if filters:
if 'estado' in filters:
conditions.append("c.estado = %s")
params.append(filters['estado'])

if 'tipo' in filters:
conditions.append("c.tipo = %s")
params.append(filters['tipo'])

if 'ubicacion' in filters:
conditions.append("l.nombre = %s")
params.append(filters['ubicacion'])

if conditions:
query += " WHERE " + " AND ".join(conditions)

query += " ORDER BY c.nombre"

# Ejecutar consulta
result = self.db.execute(query, params).fetchall()

if not result:
return {
'success': False,
'error': 'No se encontraron cámaras'
}

# Crear DataFrame
df = pd.DataFrame([{
'ID': row.id,
'Nombre': row.nombre,
'Tipo': row.tipo,
'IP': row.ip,
'Marca': row.marca,
'Modelo': row.modelo,
'Estado': row.estado,
'Fecha Instalación': row.fecha_instalacion,
'Ubicación': row.ubicacion,
'Switch Conectado': row.switch_conectado,
'IP Switch': row.switch_ip,
'Observaciones': row.observaciones
} for row in result])

# Crear archivo Excel en memoria
output = io.BytesIO()
with pd.ExcelWriter(output, engine='openpyxl') as writer:
df.to_excel(writer, sheet_name='Cámaras', index=False)

# Aplicar formato
workbook = writer.book
worksheet = writer.sheets['Cámaras']
self._format_worksheet(worksheet, df)

output.seek(0)

return {
'success': True,
'file_data': output.getvalue(),
'filename': f'reporte_camaras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
'total_records': len(df)
}

except Exception as e:
return {
'success': False,
'error': f'Error exportando cámaras: {str(e)}'
}

def export_failures_to_excel(self, date_from: str = None, date_to: str = None) -> Dict[str, Any]:
"""
Exporta fallas a Excel
"""
try:
query = """
SELECT
f.id,
f.titulo,
f.descripcion,
f.severidad,
f.estado,
f.fecha_reporte,
f.fecha_solucion,
c.nombre as camara,
u.nombre as reportado_por,
f.observaciones
FROM fallas f
LEFT JOIN camaras c ON f.camara_id = c.id
LEFT JOIN usuarios u ON f.usuario_reporta_id = u.id
"""

params = []
conditions = []

if date_from:
conditions.append("f.fecha_reporte >= %s")
params.append(date_from)

if date_to:
conditions.append("f.fecha_reporte <= %s")
params.append(date_to)

if conditions:
query += " WHERE " + " AND ".join(conditions)

query += " ORDER BY f.fecha_reporte DESC"

result = self.db.execute(query, params).fetchall()

if not result:
return {
'success': False,
'error': 'No se encontraron fallas'
}

df = pd.DataFrame([{
'ID': row.id,
'Título': row.titulo,
'Descripción': row.descripcion,
'Severidad': row.severidad,
'Estado': row.estado,
'Fecha Reporte': row.fecha_reporte,
'Fecha Solución': row.fecha_solucion,
'Cámara': row.camara,
'Reportado por': row.reportado_por,
'Observaciones': row.observaciones
} for row in result])

# Crear archivo Excel
output = io.BytesIO()
with pd.ExcelWriter(output, engine='openpyxl') as writer:
df.to_excel(writer, sheet_name='Fallas', index=False)

workbook = writer.book
worksheet = writer.sheets['Fallas']
self._format_worksheet(worksheet, df)

output.seek(0)

return {
'success': True,
'file_data': output.getvalue(),
'filename': f'reporte_fallas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
'total_records': len(df)
}

except Exception as e:
return {
'success': False,
'error': f'Error exportando fallas: {str(e)}'
}

def _format_worksheet(self, worksheet: openpyxl.worksheet, df: pd.DataFrame):
"""
Aplica formato a una hoja de Excel
"""
# Formatear encabezados
for cell in worksheet[1]:
cell.font = self.header_font
cell.fill = self.header_fill
cell.border = self.border
cell.alignment = self.center_alignment

# Ajustar ancho de columnas
for column in worksheet.columns:
max_length = 0
column_letter = column[0].column_letter

for cell in column:
try:
if len(str(cell.value)) > max_length:
max_length = len(str(cell.value))
except:
pass

adjusted_width = min(max_length + , 50)
worksheet.column_dimensions[column_letter].width = adjusted_width

# Aplicar bordes a todas las celdas con datos
for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1,
min_col=1, max_col=len(df.columns)):
for cell in row:
cell.border = self.border

# Centrar celdas con datos
for row in worksheet.iter_rows(min_row=, max_row=len(df) + 1):
for cell in row:
cell.alignment = self.center_alignment

def create_maintenance_schedule_excel(self) -> Dict[str, Any]:
"""
Crea cronograma de mantenimiento en Excel
"""
try:
# Obtener datos de mantenimiento
mantenimientos = self.db.execute("""
SELECT
m.id,
m.tipo,
m.descripcion,
m.fecha_programada,
m.fecha_ejecucion,
m.estado,
c.nombre as camara,
u.nombre as tecnico
FROM mantenimientos m
LEFT JOIN camaras c ON m.camara_id = c.id
LEFT JOIN usuarios u ON m.tecnico_id = u.id
ORDER BY m.fecha_programada
""").fetchall()

df = pd.DataFrame([{
'ID': m.id,
'Tipo': m.tipo,
'Descripción': m.descripcion,
'Fecha Programada': m.fecha_programada,
'Fecha Ejecución': m.fecha_ejecucion,
'Estado': m.estado,
'Cámara': m.camara,
'Técnico': m.tecnico
} for m in mantenimientos])

# Crear archivo Excel con múltiples hojas
output = io.BytesIO()
with pd.ExcelWriter(output, engine='openpyxl') as writer:
# Hoja principal
df.to_excel(writer, sheet_name='Cronograma', index=False)
workbook = writer.book
worksheet = writer.sheets['Cronograma']
self._format_worksheet(worksheet, df)

# Hoja de estadísticas
stats_data = self._generate_maintenance_stats()
stats_df = pd.DataFrame(list(stats_data.items()),
columns=['Métrica', 'Valor'])
stats_df.to_excel(writer, sheet_name='Estadísticas', index=False)

stats_ws = writer.sheets['Estadísticas']
self._format_worksheet(stats_ws, stats_df)

output.seek(0)

return {
'success': True,
'file_data': output.getvalue(),
'filename': f'cronograma_mantenimiento_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
'total_records': len(df)
}

except Exception as e:
return {
'success': False,
'error': f'Error creando cronograma: {str(e)}'
}

def _generate_maintenance_stats(self) -> Dict[str, Any]:
"""
Genera estadísticas de mantenimiento
"""
try:
stats = {}

# Total de mantenimientos
total = self.db.execute("SELECT COUNT(*) FROM mantenimientos").fetchone()[0]
stats['Total de Mantenimientos'] = total

# Por estado
estados = self.db.execute("""
SELECT estado, COUNT(*) as cantidad
FROM mantenimientos
GROUP BY estado
""").fetchall()

for estado in estados:
stats[f'Mantenimientos {estado.estado}'] = estado.cantidad

# Por tipo
tipos = self.db.execute("""
SELECT tipo, COUNT(*) as cantidad
FROM mantenimientos
GROUP BY tipo
""").fetchall()

for tipo in tipos:
stats[f'Mantenimientos {tipo.tipo}'] = tipo.cantidad

return stats

except Exception as e:
return {'Error': str(e)}

def import_data_from_excel(self, file_data: bytes, table_name: str) -> Dict[str, Any]:
"""
Importa datos desde Excel a una tabla de la base de datos
"""
try:
# Leer datos
df = pd.read_excel(io.BytesIO(file_data))

# Validar datos
validation = self._validate_dataframe(df)
if not validation['valid']:
return validation

# Procesar datos
processed_df = self._process_dataframe(df)

# Insertar en base de datos
inserted_count = 0
errors = []

for _, row in processed_df.iterrows():
try:
# Convertir fila a diccionario
row_dict = row.to_dict()

# Limpiar valores NaN
row_dict = {k: v for k, v in row_dict.items() if pd.notna(v)}

# Insertar según tabla
if table_name == 'camaras':
self._insert_camera(row_dict)
elif table_name == 'fallas':
self._insert_failure(row_dict)
elif table_name == 'mantenimientos':
self._insert_maintenance(row_dict)
else:
errors.append(f'Tabla {table_name} no soportada')
continue

inserted_count += 1

except Exception as e:
errors.append(f'Error insertando fila {inserted_count + 1}: {str(e)}')

self.db.commit()

return {
'success': True,
'inserted_count': inserted_count,
'errors': errors,
'total_rows': len(df)
}

except Exception as e:
self.db.rollback()
return {
'success': False,
'error': f'Error importando datos: {str(e)}'
}

def _insert_camera(self, data: Dict):
"""Inserta una cámara en la base de datos"""
query = """
INSERT INTO camaras (nombre, tipo, ip, marca, modelo, estado, fecha_instalacion, observaciones)
VALUES (%(nombre)s, %(tipo)s, %(ip)s, %(marca)s, %(modelo)s, %(estado)s, %(fecha_instalacion)s, %(observaciones)s)
"""
self.db.execute(query, data)

def _insert_failure(self, data: Dict):
"""Inserta una falla en la base de datos"""
query = """
INSERT INTO fallas (titulo, descripcion, severidad, estado, fecha_reporte, observaciones)
VALUES (%(titulo)s, %(descripcion)s, %(severidad)s, %(estado)s, %(fecha_reporte)s, %(observaciones)s)
"""
self.db.execute(query, data)

def _insert_maintenance(self, data: Dict):
"""Inserta un mantenimiento en la base de datos"""
query = """
INSERT INTO mantenimientos (tipo, descripcion, fecha_programada, estado)
VALUES (%(tipo)s, %(descripcion)s, %(fecha_programada)s, %(estado)s)
"""
self.db.execute(query, data)